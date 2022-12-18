import pandas as pd
from openpyxl import load_workbook
import openpyxl as xl
from calc_category import calcCategory, assignMarkComment


def calcCAMark(path, stream, clas, subjects, subjects_dict):
    # first open the first class
    df = pd.ExcelFile(path)

    # initialize term 1 mark
    term1Df = pd.DataFrame([], columns=subjects)
    for student in stream:
        term1Mark = 0
        subject_index = 1
        data = [[student]]
        for teacher, subject in subjects_dict.items():
            for month in ["May", "June", "October", "November"]:

                # get the marks from the different months in the different sheets
                month_sheet = df.parse(month)
                month_sheet = month_sheet.drop(
                    month_sheet.iloc[:, 0:1], axis=1)
                student_index = month_sheet["Unnamed: 3"].where(
                    month_sheet["Unnamed: 3"] == student).dropna().index.tolist()[0]
                # get the subject mark

                if month == "May":
                    mayMark = month_sheet.iloc[student_index, subject_index]
                elif month == "June":
                    juneMark = month_sheet.iloc[student_index, subject_index]
                elif month == "October":
                    octoberMark = month_sheet.iloc[student_index,
                                                   subject_index]
                elif month == "November":
                    novemberMark = month_sheet.iloc[student_index,
                                                    subject_index]

            # calculate average
            avg = float((mayMark + juneMark + octoberMark + novemberMark) / 4)

            # getting mock mark
            mockSheet = df.parse('Mock')
            mockSheet = mockSheet.drop(mockSheet.iloc[:, 0:1], axis=1)
            mockMark = mockSheet.iloc[student_index, subject_index]

            if not pd.isna(mockMark):
                caMark = ((avg + mockMark) / 2) * 0.4
                caMark = float("{:.2f}".format(caMark))
            else:
                caMark = ""
            subject_index = subject_index + 1

            # add data to list
            data[0].append(caMark)
            caMark = 0

        dataDf = pd.DataFrame(data, columns=subjects)
        term1Df = term1Df.append(dataDf, ignore_index=True)

        # dataDf.to_excel('dataDf.xlsx')
        writeDataToExcel(term1Df, 'data/average/%s.xlsx' % (clas), 'C.A.')


def writeDataToExcel(df, path, term):
    try:
        book = load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book

        df.to_excel(writer, sheet_name="%s" % (term))
        writer.close()
    except FileNotFoundError as e:
        # write data in excel file in month sheet
        df.to_excel(path, sheet_name="%s" % (term))


def calcTerm1Mark(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)

    term1Df = pd.DataFrame([], columns=subjects)
    print(len(stream))
    for student in stream:
        subject_index = 1
        data = [[student]]

        for teacher, subject in subjects_dict.items():
            for month in ["May", "June"]:

                month_sheet = df.parse(month)
                month_sheet = month_sheet.drop(
                    month_sheet.iloc[:, 0:1], axis=1)
                print('Current student: %s' % (student))
                print(month_sheet["Unnamed: 3"].where(
                    month_sheet["Unnamed: 3"] == student).dropna())
                student_index = month_sheet["Unnamed: 3"].where(
                    month_sheet["Unnamed: 3"] == student).dropna().index.tolist()[0]

                if month == "May":
                    mayMark = month_sheet.iloc[student_index, subject_index]
                elif month == "June":
                    juneMark = month_sheet.iloc[student_index, subject_index]

            avg = float((mayMark + juneMark) / 2)

            # getting mock mark
            mockSheet = df.parse('Mock')
            mockSheet = mockSheet.drop(mockSheet.iloc[:, 0:1], axis=1)
            mockMark = mockSheet.iloc[student_index, subject_index]

            if not pd.isna(mockMark):
                term1Mark = (avg + mockMark) / 2
                term1Mark = float("{:.2f}".format(term1Mark))
            else:
                term1Mark = ""

            subject_index = subject_index + 1

            # add data to list
            data[0].append(term1Mark)
            term1Mark = 0

        dataDf = pd.DataFrame(data, columns=subjects)
        term1Df = term1Df.append(dataDf, ignore_index=True)
        writeDataToExcel(term1Df, 'data/average/%s.xlsx' % (clas), "Term 1")


def calcTerm2Mark(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)

    term1Df = pd.DataFrame([], columns=subjects)

    for student in stream:
        subject_index = 1
        data = [[student]]

        for teacher, subject in subjects_dict.items():
            for month in ["October", "November"]:

                month_sheet = df.parse(month)
                month_sheet = month_sheet.drop(
                    month_sheet.iloc[:, 0:1], axis=1)
                student_index = month_sheet["Unnamed: 3"].where(
                    month_sheet["Unnamed: 3"] == student).dropna().index.tolist()[0]

                if month == "October":
                    octoberMark = month_sheet.iloc[student_index,
                                                   subject_index]
                elif month == "November":
                    novemberMark = month_sheet.iloc[student_index,
                                                    subject_index]

            avg = float((octoberMark + novemberMark) / 2)

            subject_index = subject_index + 1

            # add data to list
            data[0].append(avg)
            avg = 0

        dataDf = pd.DataFrame(data, columns=subjects)
        term1Df = term1Df.append(dataDf, ignore_index=True)
        writeDataToExcel(term1Df, 'data/average/%s.xlsx' % (clas), "Term 2")


def calcExamMark(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)

    term1Df = pd.DataFrame([], columns=subjects)

    for student in stream:
        subject_index = 1
        data = [[student]]

        for teacher, subject in subjects_dict.items():
            for month in ["Final"]:

                month_sheet = df.parse(month)
                month_sheet = month_sheet.drop(
                    month_sheet.iloc[:, 0:1], axis=1)
                student_index = month_sheet["Unnamed: 3"].where(
                    month_sheet["Unnamed: 3"] == student).dropna().index.tolist()[0]

                examMark = month_sheet.iloc[student_index, subject_index]

            avg = float(examMark * 0.6)

            subject_index = subject_index + 1

            # add data to list
            data[0].append(avg)
            avg = 0

        dataDf = pd.DataFrame(data, columns=subjects)
        term1Df = term1Df.append(dataDf, ignore_index=True)
        writeDataToExcel(term1Df, 'data/average/%s.xlsx' %
                         (clas), "Final Exam")


def calcFinalMark(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)

    finalDf = pd.DataFrame([], columns=subjects)

    for student in stream:
        subject_index = 1
        data = [[student]]

        for teacher, subject in subjects_dict.items():
            for month in ["C.A.", "Final Exam"]:

                month_sheet = df.parse(month)
                month_sheet = month_sheet.drop(
                    month_sheet.iloc[:, 0:1], axis=1)
                student_index = month_sheet["Name"].where(
                    month_sheet["Name"] == student).dropna().index.tolist()[0]

                examMark = month_sheet.iloc[student_index, subject_index]

                if month == "C.A.":
                    caMark = month_sheet.iloc[student_index, subject_index]
                elif month == "Final Exam":
                    finalExamMark = month_sheet.iloc[student_index,
                                                     subject_index]

            finalMark = float(caMark + finalExamMark)

            subject_index = subject_index + 1

            # add data to list
            data[0].append(finalMark)
            finalMark = 0

        dataDf = pd.DataFrame(data, columns=subjects)
        finalDf = finalDf.append(dataDf, ignore_index=True)
        writeDataToExcel(finalDf, 'data/average/%s.xlsx' %
                         (clas), "Final Mark")


def calcRank(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)

    sheet = df.parse('Final Mark')
    sheet = sheet.drop(sheet.iloc[:, 0:1], axis=1)

    # stud = sheet.loc[0, :].values.tolist()
    # print(type(stud))
    # print(stud)
    # agg = calc_aggregate(stud)
    # print('Aggregate %s: %.2f' % (stud[0], float(agg)))

    sheet[subjects[1]] = sheet[subjects[1]].rank(ascending=0, method='dense')
    sheet[subjects[2]] = sheet[subjects[2]].rank(ascending=0, method='dense')
    sheet[subjects[3]] = sheet[subjects[3]].rank(ascending=0, method='dense')
    sheet[subjects[4]] = sheet[subjects[4]].rank(ascending=0, method='dense')
    sheet[subjects[5]] = sheet[subjects[5]].rank(ascending=0, method='dense')
    sheet[subjects[6]] = sheet[subjects[6]].rank(ascending=0, method='dense')
    sheet[subjects[7]] = sheet[subjects[7]].rank(ascending=0, method='dense')
    sheet[subjects[8]] = sheet[subjects[8]].rank(ascending=0, method='dense')
    sheet[subjects[9]] = sheet[subjects[9]].rank(ascending=0, method='dense')
    sheet[subjects[10]] = sheet[subjects[10]].rank(ascending=0, method='dense')
    sheet[subjects[11]] = sheet[subjects[11]].rank(ascending=0, method='dense')
    sheet[subjects[12]] = sheet[subjects[12]].rank(ascending=0, method='dense')
    sheet[subjects[13]] = sheet[subjects[13]].rank(ascending=0, method='dense')
    sheet[subjects[14]] = sheet[subjects[14]].rank(ascending=0, method='dense')

    writeDataToExcel(sheet, 'data/average/%s.xlsx' % (clas), 'Positions')


def writeCategory(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)
    finalDf = pd.DataFrame([], columns=subjects)

    for student in stream:
        subject_index = 1
        data = [[student]]

        for month in ['Final Mark']:

            for teacher, subject in subjects_dict.items():
                sheet = df.parse(month)
                sheet = sheet.drop(sheet.iloc[:, 0:1], axis=1)
                student_index = sheet["Name"].where(
                    sheet["Name"] == student).dropna().index.tolist()[0]

                examMark = sheet.iloc[student_index, subject_index]

                category = calcCategory(examMark)
                examMark = ''

                if subject_index < 15:
                    subject_index = subject_index + 1
                data[0].append(category)
                category = ''

        dataDf = pd.DataFrame(data, columns=subjects)
        finalDf = finalDf.append(dataDf, ignore_index=True)
        writeDataToExcel(finalDf, 'data/average/%s.xlsx' % (clas), "Category")


def writeSubjectComments(path, stream, clas, subjects, subjects_dict):
    df = pd.ExcelFile(path)
    finalDf = pd.DataFrame([], columns=subjects)

    for student in stream:
        subject_index = 1
        data = [[student]]
        for teacher, subject in subjects_dict.items():
            sheet = df.parse('Final Mark')
            sheet = sheet.drop(sheet.iloc[:, 0:1], axis=1)
            student_index = sheet['Name'].where(
                sheet['Name'] == student).dropna().index.tolist()[0]
            examMark = sheet.iloc[student_index, subject_index]

            print('%s %s %.2f' % (student, teacher, examMark))
            comment = assignMarkComment(examMark)
            print(comment)
            examMark = ''
            data[0].append(comment)
            comment = ''
            subject_index = subject_index + 1
            # print(data)
        dataDf = pd.DataFrame(data, columns=subjects)
        finalDf = finalDf.append(dataDf, ignore_index=True)
        writeDataToExcel(finalDf, 'data/average/%s.xlsx' % (clas), "Comments")


def calc_aggregate(path, stream, minimum_aggregate, minimum_pass_mark, eng_pass_mark):
    class_average = 0
    workbook = xl.load_workbook(path)
    # print(studentMarks)
    workbook.create_sheet('Other')
    finalMark = workbook['Final Mark']
    otherSheet = workbook['Other']
    class_aggregates = []

    # assign aggregates title
    aggTitle = otherSheet.cell(1, 2)
    aggTitle.value = "Aggregate"

    # assign num passed subjects title
    cellNumPassedSubjects = otherSheet.cell(1, 4)
    cellNumPassedSubjects.value = 'Number of Passed Subjects'

    # assign english passed title
    cellEnglishPassed = otherSheet.cell(1, 5)
    cellEnglishPassed.value = 'English Passed'

    # assign term title
    termTitle = otherSheet.cell(1, 6)
    termTitle.value = "Term"

    # assign stream title
    streamTitle = otherSheet.cell(1, 7)
    streamTitle.value = 'Stream'

    for i in range(2, finalMark.max_row + 1):
        aggregate = 0
        numSubs = 0
        eng_passed = False
        num_passed_subjects = 0

        for j in range(3, finalMark.max_column + 1):
            print(finalMark.cell(i, j).value)
            item = finalMark.cell(i, j).value
            if item is None or item is str:
                continue
            elif not isinstance(float(item), float) or not isinstance(int(item), int):
                continue
            else:
                if j == 3:
                    # print('This is me')
                    if item >= eng_pass_mark:
                        eng_passed = True

                if item >= minimum_pass_mark:
                    num_passed_subjects = num_passed_subjects + 1
                aggregate = aggregate + item
                numSubs = numSubs + 1

        avg = aggregate / numSubs

        # assign average to student
        stud_avg = otherSheet.cell(i, 2)
        stud_avg.value = avg
        class_aggregates.append(avg)

        # assign num_passed_subjects to student
        numPassedSubjects = otherSheet.cell(i, 4)
        numPassedSubjects.value = num_passed_subjects

        # assign english passed variable
        englishPassed = otherSheet.cell(i, 5)
        englishPassed.value = 'PASSED' if eng_passed == True else "FAILED"

        # assign term
        term = otherSheet.cell(i, 6)
        term.value = "Term 2"

        # assign stream
        streamValue = otherSheet.cell(i, 7)
        streamValue.value = stream

    print('Total Class Aggregates: ')
    print(class_aggregates)
    print(len(class_aggregates))

    # calculate class average
    for mark in class_aggregates:
        class_average = class_average + mark

    print('Total marks: %.2f' % class_average)
    class_average = class_average / len(class_aggregates)

    # assign class aggregate to everyone
    cellClassAvgTitle = otherSheet.cell(1, 3)
    cellClassAvgTitle.value = 'Class Average'
    cellClassAvg = otherSheet.cell(2, 3)
    cellClassAvg.value = class_average

    print('This is class average: %.2f' % (class_average))

    sumMarks = 0
    totalMarks = 0
    workbook.save(path)
    workbook.close()


def calc_otherInfo(path):
    workbook = xl.load_workbook(path)
    finalMark = workbook['Final Mark']
    # workbook.create_sheet('Other')
    otherSheet = workbook['Other']

    # add class to list
    for i in range(2, finalMark.max_row + 1):
        print(finalMark.cell(i, 2).value)
        nameLiteral = otherSheet.cell(i, 1)
        nameLiteral.value = finalMark.cell(i, 2).value

    workbook.save(path)
    workbook.close()


def assignClassPosition(path):
    df = pd.ExcelFile(path)
    otherSheet = df.parse('Other')
    print(otherSheet['Aggregate'])
    otherSheet['Position'] = otherSheet['Aggregate'].rank(
        ascending=0, method='dense')
    print(otherSheet)

    writeDataToExcel(otherSheet, path, 'Other')


# def writeDataToExcel(df, path, term):
#     try:
#         book = xl.load_workbook(path)
#         writer = pd.ExcelWriter(path, engine='openpyxl')
#         writer.book = book


#         df.to_excel(writer, sheet_name="%s" % (term) )
#         writer.close()
#     except FileNotFoundError as e:
#         # write data in excel file in month sheet
#         df.to_excel(path, sheet_name="%s" % (term))
