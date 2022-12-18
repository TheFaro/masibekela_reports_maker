# class lists
import openpyxl


form1A = []
form1B = []
form1C = []
form2A = []
form2B = []
form4A = []
form4B = []

def readStudents():
    path = "data/StudentsList2022.xlsx"
    students = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    print(students.sheetnames)
    sheets = students.sheetnames
    
    # get form 1A students
    sheet = students[sheets[2]]
    # form1A = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == "NAME" :
            # print(cell.value)
            form1A.append(cell.value)
    
    # get form 1B students
    sheet = students[sheets[4]]
    # form1B = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == "Name":
            form1B.append(cell.value)
    
    # get form 1C students    
    sheet = students[sheets[6]]
    # form1C = []
    for cell in sheet['B']:
        if not cell.value is None:
            form1C.append(cell.value)
    
    # get form 2A students
    sheet = students[sheets[8]]
    # form2A = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == 'NAME':
            form2A.append(cell.value)


    # get form 2B students
    sheet = students[sheets[10]]
    # form2B = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == 'NAME':
            form2B.append(cell.value)
            
    # get form 4A students
    sheet = students[sheets[16]]
    # form4A = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == 'NAME':
            form4A.append(cell.value)
            
    # get form 4B students
    sheet = students[sheets[18]]
    # form4B = []
    for cell in sheet['B']:
        if not cell.value is None and not cell.value == 'NAME':
            form4B.append(cell.value)
