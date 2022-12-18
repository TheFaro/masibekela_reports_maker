import openpyxl as xl
import pandas as pd

class_average = 0


def calc_aggregate(path, stream, minimum_pass_mark, eng_pass_mark):
    class_average = 0
    workbook = xl.load_workbook(path)
    # print(studentMarks)
    workbook.create_sheet('Other')
    finalMark = workbook['Final Mark']
    otherSheet = workbook['Other']
    class_aggregates = []

    # assign aggregates title
    aggTitle = otherSheet.cell(1, 2)
    aggTitle.value = "Aggregate"

    # assign num passed subjects title
    cellNumPassedSubjects = otherSheet.cell(1, 4)
    cellNumPassedSubjects.value = 'Number of Passed Subjects'

    # assign english passed title
    cellEnglishPassed = otherSheet.cell(1, 5)
    cellEnglishPassed.value = 'English Passed'

    # assign term title
    termTitle = otherSheet.cell(1, 6)
    termTitle.value = "Term"

    # assign stream title
    streamTitle = otherSheet.cell(1, 7)
    streamTitle.value = 'Stream'

    for i in range(2, finalMark.max_row + 1):
        aggregate = 0
        numSubs = 0
        eng_passed = False
        num_passed_subjects = 0

        for j in range(3, finalMark.max_column + 1):
            # print(finalMark.cell(i, j).value)
            item = finalMark.cell(i, j).value
            if item is None or item is str:
                continue
            elif not isinstance(float(item), float) or not isinstance(int(item), int):
                continue
            else:
                if j == 3:
                    # print('This is me')
                    if item >= eng_pass_mark:
                        eng_passed = True

                if item >= minimum_pass_mark:
                    num_passed_subjects = num_passed_subjects + 1
                aggregate = aggregate + item
                numSubs = numSubs + 1

        avg = aggregate / numSubs

        # assign average to student
        stud_avg = otherSheet.cell(i, 2)
        stud_avg.value = avg
        class_aggregates.append(avg)

        # assign num_passed_subjects to student
        numPassedSubjects = otherSheet.cell(i, 4)
        numPassedSubjects.value = num_passed_subjects

        # assign english passed variable
        englishPassed = otherSheet.cell(i, 5)
        englishPassed.value = 'PASSED' if eng_passed == True else "FAILED"

        # assign term
        term = otherSheet.cell(i, 6)
        term.value = "Term 2"

        # assign stream
        streamValue = otherSheet.cell(i, 7)
        streamValue.value = stream

    # print('Total Class Aggregates: ')
    # print(class_aggregates)
    # print(len(class_aggregates))

    # calculate class average
    for mark in class_aggregates:
        class_average = class_average + mark

    # print('Total marks: %.2f' % class_average)
    class_average = class_average / len(class_aggregates)

    # assign class aggregate to everyone
    cellClassAvgTitle = otherSheet.cell(1, 3)
    cellClassAvgTitle.value = 'Class Average'
    cellClassAvg = otherSheet.cell(2, 3)
    cellClassAvg.value = class_average

    # print('This is class average: %.2f' % (class_average))

    workbook.save(path)
    workbook.close()


def calc_otherInfo(path):
    workbook = xl.load_workbook(path)
    finalMark = workbook['Final Mark']
    workbook.create_sheet('Other')
    otherSheet = workbook['Other']

    # add class to list
    for i in range(2, finalMark.max_row + 1):
        # print(finalMark.cell(i, 2).value)
        nameLiteral = otherSheet.cell(i, 1)
        nameLiteral.value = finalMark.cell(i, 2).value

    workbook.save(path)
    workbook.close()


def assignClassPosition(path):
    df = pd.ExcelFile(path)
    otherSheet = df.parse('Other')
    # print(otherSheet['Aggregate'])
    otherSheet['Position'] = otherSheet['Aggregate'].rank(
        ascending=0, method='dense')
    # print(otherSheet)

    writeDataToExcel(otherSheet, path, 'Other')


def writeDataToExcel(df, path, term):
    try:
        book = xl.load_workbook(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        writer.book = book

        df.to_excel(writer, sheet_name="%s" % (term))
        writer.close()
    except FileNotFoundError as e:
        # write data in excel file in month sheet
        df.to_excel(path, sheet_name="%s" % (term))

# calc_aggregate('data/average/Form1A.xlsx', 'Form 1A', 50, 50, 55)
# # print('Class average: %.2f' % (class_average))
# calc_otherInfo('data/average/Form1A.xlsx')

# assignClassPosition('data/average/Form1A.xlsx')


def getAttendance(path):

    workbook = xl.load_workbook(path, data_only=True)
    sheet = workbook['Attendance&Conduct']

    conduct = []

    for i in range(2, sheet.max_row + 1):
        stud = [sheet.cell(i, 2).value]
        for j in range(3, 6 + 1):
            # print(sheet.cell(i, j).value)
            stud.append(sheet.cell(i, j).value)
        conduct.append(stud)

    # print(conduct)
    return conduct


def makePassFail(passed_subjects, eng_passed, aggregate, level):

    if level == 'secondary':
        if passed_subjects >= 5:
            if eng_passed == 'PASSED':
                if aggregate >= 50:
                    return 'PASS'
                elif aggregate < 50:
                    return 'FAIL'
            elif eng_passed == 'FAILED':
                return 'FAIL'
        elif passed_subjects < 5:
            return 'FAIL'
    elif level == 'high school':
        if passed_subjects >= 6:
            if eng_passed == 'PASSED':
                if aggregate >= 60:
                    return 'PASS'
                elif aggregate < 60:
                    return 'FAIL'
            elif eng_passed == 'FAILED':
                return 'FAIL'
        elif passed_subjects < 6:
            return 'FAIL'


def computeClassTeacherRemarks(aggregate):
    if aggregate >= 80:
        return 'Excellent. '
    elif aggregate >= 70:
        return 'Very Good. '
    elif aggregate >= 60:
        return 'Good. '
    elif aggregate >= 50:
        return 'Fair. '
    elif aggregate < 50:
        return 'Below average. '

    # result = makePassFail(5, 'FAILED', 60, 'high school')
    # print(result)


# print(computeClassTeacherRemarks(60.3))
