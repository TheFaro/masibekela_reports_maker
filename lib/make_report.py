from __future__ import print_function

import openpyxl as xl
import class_lists

from mailmerge import MailMerge
from calc_aggregates import calc_aggregate, calc_otherInfo, assignClassPosition, getAttendance, makePassFail, computeClassTeacherRemarks


def makeClassReports(path, conduct_path, subjects, subjects_dict, stream, minimum_aggregate, minimum_pass_mark, eng_pass_mark, total_attendance_days, classTeacher, school_reopens, level):
    workbook = xl.load_workbook(path)

    # get student names
    class_lists.readStudents()
    # subjects =

    # Generating automated word document
    # template = DocxTemplate('2022 term2 report - secondary.docx')

    # Generate list of contents
    contents = []
    sheet = workbook['Term 1']
    finalData = []
    for student in range(2, sheet.max_row + 1):
        data = []
        studentName = ''

        for subject_index in range(2, sheet.max_column+1):
            subject_marks = []
            for month in ['Term 1', 'Term 2', 'C.A.', 'Final Exam', 'Final Mark', 'Positions', 'Category', 'Comments']:
                # from each sheet get the students mark
                sheet = workbook[month]

                subject_marks.append(sheet.cell(student, subject_index).value)

                if studentName == '':
                    studentName = sheet.cell(student, subject_index).value
                # print('%s %s %s' % (sheet.cell(student,2).value, month, sheet.cell(student,subject_index).value))

            data.append(subject_marks)

        finalData.append(data)

    template = "C:/Users/Faro/Documents/Work/Masibekela Reports/ReportTemplateTerm3_Form1,2,3.docx"
    document = MailMerge(template)
    # class_aggregates = []

    # Calculate subjects aggregates
    calc_aggregate(path, stream,
                   minimum_pass_mark, eng_pass_mark)

    # calc other info
    calc_otherInfo(path)

    # assign class positions
    assignClassPosition(path)

    # get other information
    otherWorkBook = xl.load_workbook(path)
    otherSheet = otherWorkBook['Other']
    othersList = []
    for i in range(2, otherSheet.max_row + 1):
        stud_dict = dict()
        for j in range(2, otherSheet.max_column + 1):
            item = otherSheet.cell(i, j).value

            if j == 2:
                stud_dict['Name'] = item
            elif j == 3:
                stud_dict['Aggregate'] = item
            elif j == 4:
                stud_dict['Class Average'] = otherSheet.cell(2, 4).value
            elif j == 5:
                stud_dict['Number of Passed Subjects'] = item
            elif j == 6:
                stud_dict['English Passed'] = item
            elif j == 7:
                stud_dict['Term'] = item
            elif j == 8:
                stud_dict['Stream'] = item
            elif j == 9:
                stud_dict['Position'] = item

        othersList.append(stud_dict)

    # print('List length')
    # print(str(float('{:.2f}'.format(othersList[0]['Class Average']))))

    conduct = getAttendance(conduct_path)
    teachers = list(subjects_dict.keys())

    for i in range(0, len(finalData)):

        # print('\n')
        data = finalData[i]

        # assign merge content
        contents.append({
            'PASSFAIL': makePassFail(othersList[i]['Number of Passed Subjects'], othersList[i]['English Passed'], othersList[i]['Aggregate'], level),
            'Class_Teacher': str(classTeacher),
            'School_Reopens': str(school_reopens),
            'Name': data[0][0],
            'Form': str(othersList[i]['Stream']),
            'Term': str(othersList[i]['Term']),
            'POSITION_in_Class': str(othersList[i]['Position']),
            'Num_of_Students': str(len(othersList)),
            'AGGREGATE': str(float('{:.2f}'.format(othersList[i]['Aggregate']))),
            'CLASS_AVERAGE': str(float('{:.2f}'.format(othersList[i]['Class Average']))),
            'Num_of_Pass': str(othersList[i]['Number of Passed Subjects']),
            'Eng_PF': str(othersList[i]['English Passed']),
            'Attendance': str(conduct[i][1]) if conduct[i][0] == data[0][0] else '',
            'Total_Days': str(total_attendance_days),
            'Conduct': str(conduct[i][4]) if conduct[i][0] == data[0][0] else '',
            'Class_teacher_s_remark': computeClassTeacherRemarks(float('{:.2f}'.format(othersList[i]['Aggregate']))),
            'Head_teachers_remark': 'Pass, good work' if float('{:.2f}'.format(data[1][4])) >= minimum_aggregate else 'Fail, work hard',
            'Staff_Resolution': 'Pass' if float('{:.2f}'.format(othersList[i]['Aggregate'])) >= minimum_aggregate else '',

            'Sub1_Content1': '' if data[1][0] is None else str(float('{:.2f}'.format(data[1][0]))),
            'Sub1_Content2': '' if data[1][1] is None else str(float('{:.2f}'.format(data[1][1]))),
            'Sub1_Content3': '' if data[1][2] is None else str(data[1][2]),
            'Sub1_Content4': '' if data[1][3] is None else str(float('{:.2f}'.format(data[1][3]))),
            'Sub1_Final_Mark': '' if data[1][4] is None else str(float('{:.2f}'.format(data[1][4]))),
            'Sub1_Position': '' if data[1][5] is None else str(data[1][5]),
            'Sub1_Category': '' if data[1][6] is None else str(data[1][6]),
            'Sub1_Comments': '' if data[1][7] is None else str(data[1][7]),
            'Sub1_Teacher': '' if data[1][5] is None else str(teachers[0]).replace('.1', ''),

            'Sub2_Content1': '' if data[2][0] is None else str(float('{:.2f}'.format(data[2][0]))),
            'Sub2_Content2': '' if data[2][1] is None else str(float('{:.2f}'.format(data[2][1]))),
            'Sub2_Content3': '' if data[2][2] is None else str(float('{:.2f}'.format(data[2][2]))),
            'Sub2_Content4': '' if data[2][3] is None else str(float('{:.2f}'.format(data[2][3]))),
            'Sub2_Final_Mark': '' if data[2][4] is None else str(float('{:.2f}'.format(data[2][4]))),
            'Sub2_Position': '' if data[2][5] is None else str(data[2][5]),
            'Sub2_Category': '' if data[2][6] is None else str(data[2][6]),
            'Sub2_Comments': '' if data[2][7] is None else str(data[2][7]),
            'Sub2_Teacher': '' if data[2][5] is None else str(teachers[1]).replace('.1', ''),

            'Sub3_Content1': '' if data[3][0] is None else str(float('{:.2f}'.format(data[3][0]))),
            'Sub3_Content2': '' if data[3][1] is None else str(float('{:.2f}'.format(data[3][1]))),
            'Sub3_Content3': '' if data[3][2] is None else str(float('{:.2f}'.format(data[3][2]))),
            'Sub3_Content4': '' if data[3][3] is None else str(float('{:.2f}'.format(data[3][3]))),
            'Sub3_Final_Mark': '' if data[3][4] is None else str(float('{:.2f}'.format(data[3][4]))),
            'Sub3_Position': '' if data[3][5] is None else str(data[3][5]),
            'Sub3_Category': '' if data[3][6] is None else str(data[3][6]),
            'Sub3_Comments': '' if data[3][7] is None else str(data[3][7]),
            'Sub3_Teacher':  '' if data[3][5] is None else str(teachers[2]).replace('.1', ''),

            'Sub4_Content1': '' if data[4][0] is None else str(float('{:.2f}'.format(data[4][0]))),
            'Sub4_Content2': '' if data[4][1] is None else str(float('{:.2f}'.format(data[4][1]))),
            'Sub4_Content3': '' if data[4][2] is None else str(float('{:.2f}'.format(data[4][2]))),
            'Sub4_Content4': '' if data[4][3] is None else str(float('{:.2f}'.format(data[4][3]))),
            'Sub4_Final_Mark': '' if data[4][4] is None else str(float('{:.2f}'.format(data[4][4]))),
            'Sub4_Position': '' if data[4][5] is None else str(data[4][5]),
            'Sub4_Category': '' if data[4][6] is None else str(data[4][6]),
            'Sub4_Comments': '' if data[4][7] is None else str(data[4][7]),
            'Sub4_Teacher': '' if data[4][5] is None else str(teachers[3]).replace('.1', ''),

            'Sub5_Content1': '' if data[5][0] is None else str(float('{:.2f}'.format(data[5][0]))),
            'Sub5_Content2': '' if data[5][1] is None else str(float('{:.2f}'.format(data[5][1]))),
            'Sub5_Content3': '' if data[5][2] is None else str(float('{:.2f}'.format(data[5][2]))),
            'Sub5_Content4': '' if data[5][3] is None else str(float('{:.2f}'.format(data[5][3]))),
            'Sub5_Final_Mark': '' if data[5][4] is None else str(float('{:.2f}'.format(data[5][4]))),
            'Sub5_Position': '' if data[5][5] is None else str(data[5][5]),
            'Sub5_Category': '' if data[5][6] is None else str(data[5][6]),
            'Sub5_Comments': '' if data[5][7] is None else str(data[5][7]),
            'Sub5_Teacher': '' if data[5][5] is None else str(teachers[4]).replace('.1', ''),

            'Sub6_Content1': '' if data[6][0] is None else str(float('{:.2f}'.format(data[6][0]))),
            'Sub6_Content2': '' if data[6][1] is None else str(float('{:.2f}'.format(data[6][1]))),
            'Sub6_Content3': '' if data[6][2] is None else str(float('{:.2f}'.format(data[6][2]))),
            'Sub6_Content4': '' if data[6][3] is None else str(float('{:.2f}'.format(data[6][3]))),
            'Sub6_Final_Mark': '' if data[6][4] is None else str(float('{:.2f}'.format(data[6][4]))),
            'Sub6_Position': '' if data[6][5] is None else str(data[6][5]),
            'Sub6_Category': '' if data[6][6] is None else str(data[6][6]),
            'Sub6_Comments': '' if data[6][7] is None else str(data[6][7]),
            'Sub6_Teacher': '' if data[6][5] is None else str(teachers[5]).replace('.1', ''),

            'Sub7_Content1': '' if data[7][0] is None else str(float('{:.2f}'.format(data[7][0]))),
            'Sub7_Content2': '' if data[7][1] is None else str(float('{:.2f}'.format(data[7][1]))),
            'Sub7_Content3': '' if data[7][2] is None else str(float('{:.2f}'.format(data[7][2]))),
            'Sub7_Content4': '' if data[7][3] is None else str(float('{:.2f}'.format(data[7][3]))),
            'Sub7_Final_Mark': '' if data[7][4] is None else str(float('{:.2f}'.format(data[7][4]))),
            'Sub7_Position': '' if data[7][5] is None else str(data[7][5]),
            'Sub7_Category': '' if data[7][6] is None else str(data[7][6]),
            'Sub7_Comments': '' if data[7][7] is None else str(data[7][7]),
            'Sub7_Teacher': '' if data[7][5] is None else str(teachers[6]).replace('.1', ''),

            'Sub8_Content1': '' if data[8][0] is None else str(float('{:.2f}'.format(data[8][0]))),
            'Sub8_Content2': '' if data[8][1] is None else str(float('{:.2f}'.format(data[8][1]))),
            'Sub8_Content3': '' if data[8][2] is None else str(float('{:.2f}'.format(data[8][2]))),
            'Sub8_Content4': '' if data[8][3] is None else str(float('{:.2f}'.format(data[8][3]))),
            'Sub8_Final_Mark': '' if data[8][4] is None else str(float('{:.2f}'.format(data[8][4]))),
            'Sub8_Position': '' if data[8][5] is None else str(data[8][5]),
            'Sub8_Category': '' if data[8][6] is None else str(data[8][6]),
            'Sub8_Comments': '' if data[8][7] is None else str(data[8][7]),
            'Sub8_Teacher': '' if data[8][5] is None else str(teachers[7]).replace('.1', ''),

            'Sub9_Content1': '' if data[9][0] is None else str(float('{:.2f}'.format(data[9][0]))),
            'Sub9_Content2': '' if data[9][1] is None else str(float('{:.2f}'.format(data[9][1]))),
            'Sub9_Content3': '' if data[9][2] is None else str(float('{:.2f}'.format(data[9][2]))),
            'Sub9_Content4': '' if data[9][3] is None else str(float('{:.2f}'.format(data[9][3]))),
            'Sub9_Final_Mark': '' if data[9][4] is None else str(float('{:.2f}'.format(data[9][4]))),
            'Sub9_Position': '' if data[9][5] is None else str(data[9][5]),
            'Sub9_Category': '' if data[9][6] is None else str(data[9][6]),
            'Sub9_Comments': '' if data[9][7] is None else str(data[9][7]),
            'Sub9_Teacher': '' if data[9][5] is None else str(teachers[8]).replace('.1', ''),

            'Sub10_Content1': '' if data[10][0] is None else str(float('{:.2f}'.format(data[10][0]))),
            'Sub10_Content2': '' if data[10][1] is None else str(float('{:.2f}'.format(data[10][1]))),
            'Sub10_Content3': '' if data[10][2] is None else str(float('{:.2f}'.format(data[10][2]))),
            'Sub10_Content4': '' if data[10][3] is None else str(float('{:.2f}'.format(data[10][3]))),
            'Sub10_Final_Mark': '' if data[10][4] is None else str(float('{:.2f}'.format(data[10][4]))),
            'Sub10_Position': '' if data[10][5] is None else str(data[10][5]),
            'Sub10_Category': '' if data[10][6] is None else str(data[10][6]),
            'Sub10_Comments': '' if data[10][7] is None else str(data[10][7]),
            'Sub10_Teacher': '' if data[10][5] is None else str(teachers[9]).replace('.1', ''),

            'Sub11_Content1': '' if data[11][0] is None else str(float('{:.2f}'.format(data[11][0]))),
            'Sub11_Content2': '' if data[11][1] is None else str(float('{:.2f}'.format(data[11][1]))),
            'Sub11_Content3': '' if data[11][2] is None else str(float('{:.2f}'.format(data[11][2]))),
            'Sub11_Content4': '' if data[11][3] is None else str(float('{:.2f}'.format(data[11][3]))),
            'Sub11_Final_Mark': '' if data[11][4] is None else str(float('{:.2f}'.format(data[11][4]))),
            'Sub11_Position': '' if data[11][5] is None else str(data[11][5]),
            'Sub11_Category': '' if data[11][6] is None else str(data[11][6]),
            'Sub11_Comments': '' if data[11][7] is None else str(data[11][7]),
            'Sub11_Teacher': '' if data[11][5] is None else str(teachers[10]).replace('.1', ''),

            'Sub12_Content1': '' if data[12][0] is None else str(float('{:.2f}'.format(data[12][0]))),
            'Sub12_Content2': '' if data[12][1] is None else str(float('{:.2f}'.format(data[12][1]))),
            'Sub12_Content3': '' if data[12][2] is None else str(float('{:.2f}'.format(data[12][2]))),
            'Sub12_Content4': '' if data[12][3] is None else str(float('{:.2f}'.format(data[12][3]))),
            'Sub12_Final_Mark': '' if data[12][4] is None else str(float('{:.2f}'.format(data[12][4]))),
            'Sub12_Position': '' if data[12][5] is None else str(data[12][5]),
            'Sub12_Category': '' if data[12][6] is None else str(data[12][6]),
            'Sub12_Comments': '' if data[12][7] is None else str(data[12][7]),
            'Sub12_Teacher': '' if data[12][5] is None else str(teachers[11]).replace('.1', ''),

            'Sub13_Content1': '' if data[13][0] is None else str(float('{:.2f}'.format(data[13][0]))),
            'Sub13_Content2': '' if data[13][1] is None else str(float('{:.2f}'.format(data[13][1]))),
            'Sub13_Content3': '' if data[13][2] is None else str(float('{:.2f}'.format(data[13][2]))),
            'Sub13_Content4': '' if data[13][3] is None else str(float('{:.2f}'.format(data[13][3]))),
            'Sub13_Final_Mark': '' if data[13][4] is None else str(float('{:.2f}'.format(data[13][4]))),
            'Sub13_Position': '' if data[13][5] is None else str(data[13][5]),
            'Sub13_Category': '' if data[13][6] is None else str(data[13][6]),
            'Sub13_Comments': '' if data[13][7] is None else str(data[13][7]),
            'Sub13_Teacher': '' if data[13][5] is None else str(teachers[12]).replace('.1', ''),

            'Sub14_Content1': '' if data[14][0] is None else str(float('{:.2f}'.format(data[14][0]))),
            'Sub14_Content2': '' if data[14][1] is None else str(float('{:.2f}'.format(data[14][1]))),
            'Sub14_Content3': '' if data[14][2] is None else str(float('{:.2f}'.format(data[14][2]))),
            'Sub14_Content4': '' if data[14][3] is None else str(float('{:.2f}'.format(data[14][3]))),
            'Sub14_Final_Mark': '' if data[14][4] is None else str(float('{:.2f}'.format(data[14][4]))),
            'Sub14_Position': '' if data[14][5] is None else str(data[14][5]),
            'Sub14_Category': '' if data[14][6] is None else str(data[14][6]),
            'Sub14_Comments': '' if data[14][7] is None else str(data[14][7]),
            'Sub14_Teacher': '' if data[14][5] is None else str(teachers[13]).replace('.1', '')
        })

    # print(contents)

    document.merge_templates(contents, separator='nextPage_section')
    document.write('%s 2022 Term 2 Report.docx' % (stream))
    document.close()


# makeClassReports('data/average/Form1A.xlsx', 'data/Attendance/Attendance_Conduct-2022-Term3-Form1A.xlsx', ['Name', 'English', 'LITinENG', 'SISWATI', 'MATH', 'INTER_SCIENCE', 'RELIGIOUS', 'ICT', 'ADD_MATHS', 'GEOGRAPHY', 'HISTORY_', 'CONSUMER_SCIENCE', 'BUSINESS_STUDIES', 'BOOK_KEEPING', 'AGRICULTURE'], {},'Form 1A', 50, 50, 55,70)
